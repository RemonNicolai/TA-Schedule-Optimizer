import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

# name shopkeepers and select days
shopkeepers = ["Remon", "Sebastian", "Aya", "Jadwiga"]
block_number = "1b"
first_day = datetime.strptime('10/11/2024', '%d/%m/%Y')  # pick the Sunday before
last_day = datetime.strptime('02/02/2025', '%d/%m/%Y')

# Create CSV file to specify availability with shopkeepers
def create_availability_excel(shopkeepers, first_day, last_day, block_number = str()):

    cd = os.getcwd()
    output_path = os.path.join(cd,'availability-template output')
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    filename = "availability" + block_number + ".xlsx"
    full_path = os.path.join(output_path, filename)
    day_list = []
    current_day = first_day

    while current_day <= last_day:
        if current_day.weekday() < 5:  # Only weekdays (Monday to Friday)
            day_list.append(current_day)
        current_day += timedelta(days=1)

    day_list_dup = [day for day in day_list for _ in range(2)]

    schedule_array = []
    for i, day in enumerate(day_list_dup):
        day_name = day.strftime('%A')
        day_date = day.strftime('%d/%m/%Y')
        time_slot = '09:00' if i % 2 == 0 else '13:00'
        schedule_array.append([day_name, day_date, time_slot])

    # Create DataFrame with Day, Date, Time columns
    schedule_df = pd.DataFrame(schedule_array, columns=['Day', 'Date', 'Time'])

    # Add columns for each shopkeeper with empty values
    for shopkeeper in shopkeepers:
        schedule_df[shopkeeper] = ''

    # Pre-fill first 3 rows of the first shopkeeper's column with examples
    schedule_df.iloc[0, 3] = 'Yes'
    schedule_df.iloc[1, 3] = 'Preferably Not'
    schedule_df.iloc[2, 3] = 'No'

    # Save DataFrame to an Excel file (CSV doesn't support data validation)
    schedule_df.to_excel(full_path, index=False)

    # Load the workbook and worksheet for adding validation and conditional formatting
    wb = load_workbook(full_path)
    ws = wb.active

    # Apply validation to the shopkeeper columns (starting from column 4)
    for col in range(4, 4 + len(shopkeepers)):
        dv = DataValidation(type="list", formula1='"Yes,Preferably Not,No"', showDropDown=False)

        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'

        ws.add_data_validation(dv)

        # Get the column letter (e.g., 'D', 'E', etc.)
        col_letter = ws.cell(row=1, column=col).column_letter
        cell_range = f"{col_letter}2:{col_letter}{len(schedule_df) + 1}"

        # Add the data validation to the cell range
        dv.add(cell_range)

    # Create conditional formatting rules
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    range_end = f"H{len(schedule_df) + 1}"
    range_string = "D2:" + range_end

    colour_dict = {'Yes': green_fill, 'Preferably Not': yellow_fill, 'No': red_fill}

    # apply conditional formatting rules based on the dictionary
    for key, fill in colour_dict.items():
        ws.conditional_formatting.add(range_string,
                                      Rule(type='containsText', text = str(key), stopIfTrue=True,
                                           dxf=DifferentialStyle(fill=fill)))

    # Save the Excel file with data validation and conditional formatting
    wb.save(full_path)
    print(
        f"Availability excel sheet created successfully and written to {full_path}")

    RED_TEXT = "\033[91m"
    RESET_TEXT = "\033[0m"
    print(
        f"{RED_TEXT}Don't forget to manually check for breaks and other holidays.{RESET_TEXT}")

create_availability_excel(shopkeepers, first_day, last_day, block_number)